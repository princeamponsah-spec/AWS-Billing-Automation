"""
Writes one week's worth of figures (6 rows, one per account) into the existing
'Weekly Entry' tab of AWS_Weekly_Billing_Tracker.xlsx, WITHOUT touching any formulas.

Design notes:
- The tracker is pre-built with a fixed number of weekly 6-row blocks (see TOTAL_WEEKS
  in the original build script). Only columns A-I are ever blank-and-waiting for each
  future week; columns J onward already contain formulas that will calculate correctly
  the moment A-I are filled in. This module therefore only ever writes to A-I.
- If every pre-built block is already full, this module raises TrackerFullError instead
  of guessing how to extend the formula ranges - see the README for how to add more
  weeks by hand (or as a future enhancement).
"""

from openpyxl import load_workbook

ACCOUNTS_PER_WEEK = 6
HEADER_ROW = 1
SHEET_NAME = "Weekly Entry"


class TrackerFullError(Exception):
    """Raised when there is no pre-built blank week block left to write into."""


def find_next_empty_block(ws) -> int:
    """Return the row number where the next empty 6-row week block starts, or raise
    TrackerFullError if the sheet has no remaining blank blocks."""
    row = HEADER_ROW + 1
    while True:
        cell = ws.cell(row=row, column=1)
        if cell.value in (None, ""):
            return row
        row += ACCOUNTS_PER_WEEK
        if row > ws.max_row + ACCOUNTS_PER_WEEK:
            # Safety valve - shouldn't normally trigger since a truly full sheet is
            # caught by the max_row check in write_week below.
            raise TrackerFullError(
                "No blank week block found within the sheet's built range."
            )


def write_week(workbook_path: str, week_ending, account_order: list, figures_by_account: dict):
    """
    week_ending: datetime.date - the Sunday this report week ends on.
    account_order: list of account-name strings, in the exact order the sheet expects.
    figures_by_account: {account_name: {mtd, last_month_same_period, forecast,
                                         last_month_total, highest_spend,
                                         highest_service, trend}}

    Saves the workbook in place (caller is responsible for the S3 download/upload
    round-trip - see lambda_function.py).
    """
    wb = load_workbook(workbook_path)
    ws = wb[SHEET_NAME]

    start_row = find_next_empty_block(ws)
    end_row = start_row + ACCOUNTS_PER_WEEK - 1
    if end_row > ws.max_row:
        raise TrackerFullError(
            f"Writing rows {start_row}-{end_row} would exceed the sheet's built range "
            f"(max row {ws.max_row}). Extend the tracker with more pre-built week blocks "
            f"(and their formulas) before the next run - see README.md."
        )

    for offset, account in enumerate(account_order):
        row = start_row + offset
        figures = figures_by_account.get(account)
        if figures is None:
            raise ValueError(f"No figures supplied for account '{account}'")

        ws.cell(row=row, column=1, value=week_ending)                       # A Week Ending
        ws.cell(row=row, column=2, value=account)                           # B AWS Account
        ws.cell(row=row, column=3, value=figures["mtd"])                    # C MTD Cost
        ws.cell(row=row, column=4, value=figures["last_month_same_period"]) # D Last Month Same Period
        ws.cell(row=row, column=5, value=figures["forecast"])               # E Forecast
        ws.cell(row=row, column=6, value=figures["last_month_total"])       # F Last Month Total
        ws.cell(row=row, column=7, value=figures["highest_spend"])          # G Highest Service Spend
        ws.cell(row=row, column=8, value=figures["highest_service"])        # H Highest Spend Service
        if figures["trend"] is not None:
            ws.cell(row=row, column=9, value=figures["trend"])              # I Trend vs Prior Month

    wb.save(workbook_path)
    return start_row, end_row
